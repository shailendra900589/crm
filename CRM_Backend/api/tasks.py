import io







from celery import shared_task



from django.contrib.auth import get_user_model







User = get_user_model()











def _broadcast(event):



    try:



        from asgiref.sync import async_to_sync



        from channels.layers import get_channel_layer







        layer = get_channel_layer()



        if not layer:



            return



        for group in ("dashboard_admin", "dashboard_bdm", "dashboard_manager"):



            async_to_sync(layer.group_send)(group, event)



    except Exception:



        pass











@shared_task(bind=True)



def process_bulk_upload_task(self, job_id):



    process_bulk_upload(job_id)











def process_bulk_upload(job_id):



    from openpyxl import load_workbook







    from .models import BulkUploadJob, CustomForm, Lead, Merchant







    job = BulkUploadJob.objects.select_related("project").get(id=job_id)



    job.status = BulkUploadJob.Status.PROCESSING



    job.save(update_fields=["status"])



    _broadcast({"type": "dashboard_update", "model": "BulkUploadJob", "id": job.id, "job_id": job.id})







    errors = []



    success = 0



    custom_form = CustomForm.objects.filter(project=job.project, is_active=True).first()



    schema = custom_form.schema if custom_form else []







    try:



        wb = load_workbook(job.file.path, read_only=True)



        ws = wb.active



        rows = list(ws.iter_rows(values_only=True))



        if not rows:



            raise ValueError("Empty spreadsheet")



        headers = [str(h).strip().lower() if h else "" for h in rows[0]]



        job.total_rows = len(rows) - 1



        job.save(update_fields=["total_rows"])







        for i, row in enumerate(rows[1:], start=2):



            data = dict(zip(headers, row))



            try:



                name = str(data.get("merchant_name") or data.get("name") or "").strip()



                mobile = str(data.get("mobile") or data.get("merchant_mobile") or "").strip()



                if not name or not mobile:



                    raise ValueError("merchant_name and mobile required")







                custom_data = {}



                for field in schema:



                    fid = (field.get("field_id") or field.get("label", "")).lower()



                    label = str(field.get("label", "")).lower()



                    val = data.get(fid) or data.get(label)



                    if field.get("required") and (val is None or str(val).strip() == ""):



                        raise ValueError(f"Missing required field: {field.get('label')}")



                    if val is not None and str(val).strip():



                        custom_data[field.get("field_id") or fid] = val







                merchant, _ = Merchant.objects.update_or_create(



                    project=job.project,



                    mobile=mobile,



                    defaults={



                        "name": name,



                        "email": str(data.get("email") or ""),



                        "brand_name": str(data.get("brand_name") or ""),



                        "city": str(data.get("city") or ""),



                    },



                )



                if Lead.objects.filter(project=job.project, merchant=merchant).exists():



                    errors.append({"row": i, "error": f"Duplicate skipped — lead already exists for {mobile}"})



                    continue



                Lead.objects.create(



                    project=job.project,



                    merchant=merchant,



                    bdm=job.uploaded_by,



                    status=str(data.get("status") or "interested"),



                    notes=str(data.get("notes") or ""),



                    custom_data=custom_data,



                )



                success += 1



                job.success_rows = success



                if i % 10 == 0:



                    job.save(update_fields=["success_rows"])



                    _broadcast({



                        "type": "dashboard_update",



                        "model": "BulkUploadJob",



                        "id": job.id,



                        "job_id": job.id,



                        "progress": success,



                        "total": job.total_rows,



                    })



            except Exception as e:



                errors.append({"row": i, "error": str(e)})







        job.status = BulkUploadJob.Status.DONE



        job.success_rows = success



        job.error_log = errors



        job.save()



    except Exception as e:



        job.status = BulkUploadJob.Status.FAILED



        job.error_log = [{"error": str(e)}]



        job.save()







    _broadcast({"type": "dashboard_update", "model": "BulkUploadJob", "id": job.id, "job_id": job.id})



    return job.id











def generate_bulk_template(project):



    from openpyxl import Workbook



    from openpyxl.worksheet.datavalidation import DataValidation







    wb = Workbook()



    ws = wb.active



    ws.title = "Leads"



    headers = ["merchant_name", "mobile", "email", "brand_name", "city", "status", "notes"]

    from .models import CustomForm

    form = CustomForm.objects.filter(project=project, is_active=True).first()



    if form and form.schema:



        for field in form.schema:



            headers.append(field.get("field_id") or field.get("label", "field"))



    ws.append(headers)



    example = ["Demo Store", "9999999999", "demo@store.com", "Demo Brand", "Mumbai", "interested", "Sample row"]



    if form and form.schema:



        for field in form.schema:



            if field.get("type") == "dropdown" and field.get("options"):



                example.append(field["options"][0])



            else:



                example.append("")



    ws.append(example)







    if form and form.schema:



        for idx, field in enumerate(form.schema):



            if field.get("type") == "dropdown" and field.get("options"):



                col = len(["merchant_name", "mobile", "email", "brand_name", "city", "status", "notes"]) + idx + 1



                col_letter = chr(64 + col) if col <= 26 else "A"



                opts = ",".join(field["options"])



                dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=not field.get("required"))



                dv.add(f"{col_letter}2:{col_letter}500")



                ws.add_data_validation(dv)







    buf = io.BytesIO()



    wb.save(buf)



    buf.seek(0)



    return buf

from .digest import send_daily_digest


@shared_task(name="api.tasks.send_daily_digest_task")
def send_daily_digest_task(user_id=None):
    return send_daily_digest(user_id=user_id)
