import io
from datetime import date

from django.http import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _style_header(ws, cols):
    fill = PatternFill("solid", fgColor="1E3A8A")
    font = Font(color="FFFFFF", bold=True)
    for i, title in enumerate(cols, 1):
        cell = ws.cell(1, i, title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 40)


def _answer_display(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return value.get("url") or value.get("name") or str(value)
    return str(value)


def _form_columns_for_leads(leads):
    """Ordered unique form field columns from project forms + answers."""
    from .models import CustomForm

    project_ids = {getattr(l, "project_id", None) for l in leads}
    project_ids.discard(None)
    forms = {
        f.project_id: f
        for f in CustomForm.objects.filter(project_id__in=project_ids or [-1]).only("project_id", "schema", "title")
    }

    columns = []  # list of (field_id, label)
    seen = set()
    for pid in sorted(project_ids):
        form = forms.get(pid)
        if not form:
            continue
        for field in form.schema or []:
            if not isinstance(field, dict):
                continue
            fid = field.get("field_id")
            if not fid or fid in seen:
                continue
            seen.add(fid)
            columns.append((fid, field.get("label") or fid))

    # Include any answer keys not in schema (legacy fills)
    for lead in leads:
        data = lead.custom_data or {}
        if not isinstance(data, dict):
            continue
        for fid in data.keys():
            if fid in seen:
                continue
            seen.add(fid)
            columns.append((fid, fid))
    return columns


def export_leads_xlsx(leads, filename="leads_export.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    form_cols = _form_columns_for_leads(leads)
    headers = [
        "ID", "Project", "Product", "Merchant", "Mobile", "Email", "Brand", "City",
        "BDM", "Status", "Follow-up Date", "Notes", "Created At", "Updated At",
    ] + [label for _, label in form_cols]
    _style_header(ws, headers)

    for row, lead in enumerate(leads, 2):
        ws.cell(row, 1, lead.id)
        ws.cell(row, 2, lead.project.name if lead.project_id else "")
        ws.cell(row, 3, lead.product.name if lead.product_id else "")
        ws.cell(row, 4, lead.merchant.name)
        ws.cell(row, 5, lead.merchant.mobile)
        ws.cell(row, 6, lead.merchant.email or "")
        ws.cell(row, 7, lead.merchant.brand_name or "")
        ws.cell(row, 8, lead.merchant.city or "")
        ws.cell(row, 9, lead.bdm.get_full_name() or lead.bdm.username)
        ws.cell(row, 10, lead.get_status_display())
        ws.cell(row, 11, str(lead.follow_up_date) if lead.follow_up_date else "")
        ws.cell(row, 12, lead.notes or "")
        ws.cell(row, 13, lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else "")
        ws.cell(row, 14, lead.updated_at.strftime("%Y-%m-%d %H:%M") if lead.updated_at else "")
        data = lead.custom_data if isinstance(lead.custom_data, dict) else {}
        for i, (fid, _) in enumerate(form_cols):
            ws.cell(row, 15 + i, _answer_display(data.get(fid)))

    _autosize(ws)

    # Dedicated sheet mirroring form field order / BDM answers
    ws2 = wb.create_sheet("Form Answers")
    form_headers = [
        "Lead ID", "Project", "Merchant", "BDM", "Status", "Submitted At",
    ] + [label for _, label in form_cols]
    _style_header(ws2, form_headers)
    r = 2
    for lead in leads:
        data = lead.custom_data if isinstance(lead.custom_data, dict) else {}
        if not data and not getattr(lead, "form_submissions", None):
            continue
        sub = None
        try:
            sub = lead.form_submissions.order_by("-submitted_at").first()
        except Exception:
            sub = None
        answers = (sub.answers if sub and isinstance(sub.answers, dict) else None) or data
        if not answers:
            continue
        ws2.cell(r, 1, lead.id)
        ws2.cell(r, 2, lead.project.name if lead.project_id else "")
        ws2.cell(r, 3, lead.merchant.name)
        ws2.cell(r, 4, lead.bdm.get_full_name() or lead.bdm.username)
        ws2.cell(r, 5, lead.get_status_display())
        ws2.cell(
            r,
            6,
            sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub and sub.submitted_at else (
                lead.updated_at.strftime("%Y-%m-%d %H:%M") if lead.updated_at else ""
            ),
        )
        for i, (fid, _) in enumerate(form_cols):
            ws2.cell(r, 7 + i, _answer_display(answers.get(fid)))
        r += 1
    _autosize(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FileResponse(
        buf,
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def export_admin_report_xlsx(payload, filename=None, leads=None):
    wb = Workbook()
    leads = leads or []

    # Summary
    ws = wb.active
    ws.title = "Summary"
    _style_header(ws, ["Metric", "Value"])
    money = payload.get("money_metrics") or {}
    summary_rows = [
        ("Total Projects", payload.get("total_projects", 0)),
        ("Active Projects", payload.get("active_projects", 0)),
        ("Companies", payload.get("total_companies", 0)),
        ("Products", payload.get("total_products", 0)),
        ("Total Leads", payload.get("total_leads", 0)),
        ("Orders Confirmed", payload.get("orders_confirmed", 0)),
        ("Conversion Rate %", payload.get("conversion_rate", 0)),
        ("Follow-ups Today", payload.get("follow_ups_due_today", 0)),
        ("Overdue Follow-ups", payload.get("overdue_follow_ups", 0)),
        ("Forms Filled Today", payload.get("forms_filled_today", 0)),
        ("Visits Today", payload.get("visits_scheduled_today", 0)),
        ("Amount Collected", money.get("total_collection", 0)),
        ("Collection Pending", money.get("total_pending", 0)),
    ]
    for i, (k, v) in enumerate(summary_rows, 2):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)
    _autosize(ws)

    # Projects
    ws = wb.create_sheet("Projects")
    _style_header(ws, ["Project", "Leads", "Confirmed", "Conversion %", "Status"])
    for i, p in enumerate(payload.get("project_stats") or [], 2):
        ws.cell(i, 1, p.get("name"))
        ws.cell(i, 2, p.get("lead_count", 0))
        ws.cell(i, 3, p.get("confirmed_count", 0))
        ws.cell(i, 4, p.get("conversion", 0))
        ws.cell(i, 5, "Active" if p.get("is_active") else "Inactive")
    _autosize(ws)

    # Companies
    ws = wb.create_sheet("Companies")
    _style_header(ws, ["Company", "City", "Project", "Leads", "Confirmed", "Conversion %"])
    for i, c in enumerate(payload.get("company_stats") or [], 2):
        ws.cell(i, 1, c.get("name"))
        ws.cell(i, 2, c.get("city") or "")
        ws.cell(i, 3, c.get("project_name") or "")
        ws.cell(i, 4, c.get("lead_count", 0))
        ws.cell(i, 5, c.get("confirmed_count", 0))
        ws.cell(i, 6, c.get("conversion", 0))
    _autosize(ws)

    # Products
    ws = wb.create_sheet("Products")
    _style_header(ws, ["Product", "Project", "Leads", "Confirmed", "Conversion %"])
    for i, p in enumerate(payload.get("product_stats") or [], 2):
        ws.cell(i, 1, p.get("name"))
        ws.cell(i, 2, p.get("project_name") or "")
        ws.cell(i, 3, p.get("lead_count", 0))
        ws.cell(i, 4, p.get("confirmed_count", 0))
        ws.cell(i, 5, p.get("conversion", 0))
    _autosize(ws)

    # Disposition
    ws = wb.create_sheet("Disposition")
    _style_header(ws, ["Status", "Count"])
    for i, d in enumerate(payload.get("disposition") or [], 2):
        ws.cell(i, 1, d.get("status"))
        ws.cell(i, 2, d.get("count", 0))
    _autosize(ws)

    # Form answers — exact BDM fill matching form field labels
    if leads:
        form_cols = _form_columns_for_leads(leads)
        ws = wb.create_sheet("Form Submissions")
        headers = [
            "Lead ID", "Project", "Merchant", "City", "BDM", "Product", "Status", "Submitted At",
        ] + [label for _, label in form_cols]
        _style_header(ws, headers)
        r = 2
        for lead in leads:
            data = lead.custom_data if isinstance(lead.custom_data, dict) else {}
            sub = None
            try:
                subs = list(lead.form_submissions.all())
                if subs:
                    sub = max(
                        subs,
                        key=lambda s: s.submitted_at.timestamp() if s.submitted_at else 0,
                    )
            except Exception:
                sub = None
            answers = (sub.answers if sub and isinstance(sub.answers, dict) else None) or data
            if not answers:
                continue
            ws.cell(r, 1, lead.id)
            ws.cell(r, 2, lead.project.name if lead.project_id else "")
            ws.cell(r, 3, lead.merchant.name)
            ws.cell(r, 4, lead.merchant.city or "")
            ws.cell(r, 5, lead.bdm.get_full_name() or lead.bdm.username)
            ws.cell(r, 6, lead.product.name if lead.product_id else "")
            ws.cell(r, 7, lead.get_status_display())
            ws.cell(
                r,
                8,
                sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub and sub.submitted_at else (
                    lead.updated_at.strftime("%Y-%m-%d %H:%M") if lead.updated_at else ""
                ),
            )
            for i, (fid, _) in enumerate(form_cols):
                ws.cell(r, 9 + i, _answer_display(answers.get(fid)))
            r += 1
        _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = filename or f"crm_report_{date.today().isoformat()}.xlsx"
    return FileResponse(
        buf,
        as_attachment=True,
        filename=name,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def export_admin_report_pdf(payload, filename=None, leads=None):
    from fpdf import FPDF

    class ReportPDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 11)
            self.set_text_color(30, 64, 175)
            self.cell(0, 8, "Sales CRM - Performance Report", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "", 8)
            self.set_text_color(120, 120, 120)
            self.cell(0, 5, f"Generated {date.today().isoformat()}", align="C", new_x="LMARGIN", new_y="NEXT")
            self.ln(3)

        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 8, f"Page {self.page_no()}", align="C")

        def section(self, title):
            self.ln(4)
            self.set_font("Helvetica", "B", 12)
            self.set_text_color(15, 23, 42)
            self.cell(0, 8, title, new_x="LMARGIN", new_y="NEXT")
            self.set_draw_color(226, 232, 240)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(3)

        def kv(self, label, value):
            self.set_font("Helvetica", "", 9)
            self.set_text_color(71, 85, 105)
            self.cell(70, 6, str(label))
            self.set_text_color(15, 23, 42)
            self.set_font("Helvetica", "B", 9)
            self.cell(0, 6, str(value), new_x="LMARGIN", new_y="NEXT")

        def table(self, headers, rows, widths):
            self.set_font("Helvetica", "B", 8)
            self.set_fill_color(30, 58, 138)
            self.set_text_color(255, 255, 255)
            for h, w in zip(headers, widths):
                self.cell(w, 7, str(h), border=0, fill=True)
            self.ln()
            self.set_font("Helvetica", "", 8)
            self.set_text_color(30, 41, 59)
            fill = False
            for row in rows:
                if self.get_y() > 270:
                    self.add_page()
                self.set_fill_color(248, 250, 252)
                for val, w in zip(row, widths):
                    self.cell(w, 6, str(val)[:40], border=0, fill=fill)
                self.ln()
                fill = not fill

    pdf = ReportPDF()
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_page()

    summary = payload.get("filter_summary") or {}
    if summary.get("project_name") or summary.get("from") or summary.get("to"):
        pdf.section("Active Filters")
        if summary.get("project_name"):
            pdf.kv("Project", summary["project_name"])
        if summary.get("from") or summary.get("to"):
            pdf.kv("Date range", f"{summary.get('from') or '...'} to {summary.get('to') or '...'}")

    pdf.section("KPI Summary")
    money = payload.get("money_metrics") or {}
    for label, key in [
        ("Total Projects", "total_projects"),
        ("Companies", "total_companies"),
        ("Products", "total_products"),
        ("Total Leads", "total_leads"),
        ("Orders Confirmed", "orders_confirmed"),
        ("Conversion Rate %", "conversion_rate"),
        ("Follow-ups Today", "follow_ups_due_today"),
        ("Overdue Follow-ups", "overdue_follow_ups"),
        ("Forms Filled Today", "forms_filled_today"),
        ("Visits Today", "visits_scheduled_today"),
    ]:
        pdf.kv(label, payload.get(key, 0))
    pdf.kv("Amount Collected", money.get("total_collection", 0))
    pdf.kv("Collection Pending", money.get("total_pending", 0))

    projects = payload.get("project_stats") or []
    if projects:
        pdf.section("Projects")
        pdf.table(
            ["Project", "Leads", "Confirmed", "Conv %", "Status"],
            [
                [
                    p.get("name", ""),
                    p.get("lead_count", 0),
                    p.get("confirmed_count", 0),
                    p.get("conversion", 0),
                    "Active" if p.get("is_active") else "Inactive",
                ]
                for p in projects[:40]
            ],
            [55, 30, 35, 30, 30],
        )

    companies = payload.get("company_stats") or []
    if companies:
        pdf.section("Companies")
        pdf.table(
            ["Company", "City", "Leads", "Confirmed", "Conv %"],
            [
                [
                    c.get("name", ""),
                    c.get("city") or "-",
                    c.get("lead_count", 0),
                    c.get("confirmed_count", 0),
                    c.get("conversion", 0),
                ]
                for c in companies[:40]
            ],
            [55, 35, 30, 35, 25],
        )

    products = payload.get("product_stats") or []
    if products:
        pdf.section("Products")
        pdf.table(
            ["Product", "Project", "Leads", "Confirmed", "Conv %"],
            [
                [
                    p.get("name", ""),
                    p.get("project_name") or "-",
                    p.get("lead_count", 0),
                    p.get("confirmed_count", 0),
                    p.get("conversion", 0),
                ]
                for p in products[:40]
            ],
            [50, 45, 30, 35, 20],
        )

    disposition = payload.get("disposition") or []
    if disposition:
        pdf.section("Lead Disposition")
        pdf.table(
            ["Status", "Count"],
            [[d.get("status", ""), d.get("count", 0)] for d in disposition],
            [100, 40],
        )

    # Compact form submissions preview (labels + first few answers)
    if leads:
        filled = [l for l in leads if isinstance(l.custom_data, dict) and l.custom_data][:25]
        if filled:
            form_cols = _form_columns_for_leads(filled)[:4]
            pdf.add_page()
            pdf.section("Form Submissions (sample)")
            headers = ["Merchant", "BDM"] + [lab for _, lab in form_cols]
            widths = [50, 35] + [28] * len(form_cols)
            rows = []
            for lead in filled:
                answers = lead.custom_data or {}
                rows.append(
                    [lead.merchant.name, lead.bdm.get_full_name() or lead.bdm.username]
                    + [_answer_display(answers.get(fid))[:28] for fid, _ in form_cols]
                )
            pdf.table(headers, rows, widths)

    out = pdf.output()
    buf = io.BytesIO(out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1"))
    buf.seek(0)
    name = filename or f"crm_report_{date.today().isoformat()}.pdf"
    return FileResponse(buf, as_attachment=True, filename=name, content_type="application/pdf")


def export_leads_pdf(leads, filename="leads_export.pdf"):
    from fpdf import FPDF

    rows = list(leads)[:500]
    form_cols = _form_columns_for_leads(rows)[:5]
    pdf = FPDF(orientation="L", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(30, 64, 175)
    pdf.cell(0, 8, "Leads Export", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, f"Generated {date.today().isoformat()} · {len(rows)} records", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    headers = ["ID", "Project", "Merchant", "City", "BDM", "Status"] + [lab for _, lab in form_cols]
    base_w = [14, 28, 40, 28, 30, 32]
    form_w = [max(22, min(36, 180 // max(len(form_cols), 1)))] * len(form_cols)
    widths = base_w + form_w
    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(30, 58, 138)
    pdf.set_text_color(255, 255, 255)
    for h, w in zip(headers, widths):
        pdf.cell(w, 6, str(h)[:18], fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(30, 41, 59)
    fill = False
    for lead in rows:
        if pdf.get_y() > 190:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 7)
            pdf.set_fill_color(30, 58, 138)
            pdf.set_text_color(255, 255, 255)
            for h, w in zip(headers, widths):
                pdf.cell(w, 6, str(h)[:18], fill=True)
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            pdf.set_text_color(30, 41, 59)
        pdf.set_fill_color(248, 250, 252)
        data = lead.custom_data if isinstance(lead.custom_data, dict) else {}
        row = [
            lead.id,
            lead.project.name if lead.project_id else "",
            lead.merchant.name,
            lead.merchant.city or "",
            lead.bdm.get_full_name() or lead.bdm.username,
            lead.get_status_display(),
        ] + [_answer_display(data.get(fid)) for fid, _ in form_cols]
        for val, w in zip(row, widths):
            pdf.cell(w, 5.5, str(val)[:28], fill=fill)
        pdf.ln()
        fill = not fill

    out = pdf.output()
    buf = io.BytesIO(out if isinstance(out, (bytes, bytearray)) else out.encode("latin-1"))
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=filename, content_type="application/pdf")
